import os
from  openpyxl import*
from  openpyxl.comments import Comment
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
path='data.xlsx'
if os.path.exists(path):
    print('File already exists')
    workbook=load_workbook(path)
    ws =workbook.active
    t= ws.tables['t1']
else:
    workbook=Workbook()
    workbook[workbook.sheetnames[0]].title='Data1'
    ws =workbook.active
    if 't1' not in ws.tables:
        ws.append(['Name','Age','Email','Gender'])
        gv = DataValidation(type='list',formula1='"Male,Female"',allow_blank=True)
        ws.add_data_validation(gv)
        gv.add('D2:D1048576')
        t=Table(displayName='t1',ref=f'A1:{get_column_letter(ws.max_column)}{ws.max_row}')
        ws.add_table(t)
    else:t=ws.tables['t1']
# print(ws.max_column,ws.max_row)prints the no of col n row
# ws.delete_rows(1,4)
# ws.delete_cols(1,4)
# ws.append(['Id','Name','Age','Languages Known'])
ws.append(['A',23,'A@gmail.com'])
ws.append(['B',24,'B@gmail.com'])
ws.append(['C',24,'C@gmail.com'])
# ws.delete_rows(5,3)
t.ref = f'A1:{get_column_letter(ws.max_column)}{ws.max_row}'
# workbook.create_sheet('Data2')
# print(workbook.sheetnames)
for i in ws.iter_rows():
    for j in i:
        print(j.value,end=" ")
    print()
ws['F3'].comment=Comment(text='This is a comment',author='Naveen')
workbook.save(path)
workbook.close()
